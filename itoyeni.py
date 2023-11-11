from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook,load_workbook
from helium import *
import sqlite3
dizi=[]
connection = sqlite3.connect("db.sqlite3")
cursor = connection.cursor()
wb = load_workbook("A.xlsx")
ws = wb.active
##firma_no=int(ws["D1"].value)


    
firma_no=int(input("Firma noyu giriniz :"))    

driver = start_chrome("https://bilgibankasi.ito.org.tr/tr/bilgi-bankasi/firma-bilgileri")
def firma_cek(company_code):
    gecilen=0
    eklenen=0    
    
    time.sleep(2)

##    baha = driver.find_element(By.XPATH, "/html/body/div[3]/div/div/div[2]/div/div/div/div/div/div/div/div/div[1]/div/div[1]/div/div/div/div/div/div/form/div/div[1]/div/div[1]/div/label/input")
    doubleclick("Ticaret Sicil No")
    press(DELETE)
 
    write(company_code,into="Ticaret Sicil No")
##    write(company_code)
    ##    baha.send_keys(company_code)
    click("ara")
    time.sleep(3)
    try:
        secici="/html/body/div[3]/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td[2]"
        firma=driver.find_element(By.XPATH,secici)
        click(firma)

    except NoSuchElementException:
         ##         baha=driver.find_element(By.XPATH,"/html/body/div[3]/div/div/div[2]/div /div/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[1]/td[2]")
    
         secici="/html/body/div[3]/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[1]/td[2]"
         firma=driver.find_element(By.CSS_SELECTOR,secici)
         click(firma)
#baha=driver.find_element(By.XPATH,"/html/body/div[3]/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td[2]")
    
    time.sleep(2)
    elements=driver.find_elements(By.CLASS_NAME,"data-cell.col-md-9")

    for element in elements:
        dizi.append(element.text)
##    ws.append(dizi)
    for index, element10 in enumerate(dizi, start=1):
        print(f"{index}. eleman: {element10}")
    
##    ws["D1"].value=str(company_code)
##    wb.save("A.xlsx")
    sector=dizi[13]
    name=dizi[4]
    phone=dizi[6]
    if phone== "" : phone=company_code
            
    site=dizi[8]
    if site=="" : site=" "
    address=dizi[5]
    personels_caount=dizi[1]
    note="--kuruluş :"+dizi[9]+"--sermaye :"+dizi[12]+"--açıklama :"

    created_date="2023-08-20 23:15:54.575094"
    updated_date="2023-08-20 23:15:54.575094"
    city_id="17"
    fount_id="1"
    last_status_id="1"
    user_id="1"
    full_name=" "
    reminder=" "
    short_name1=dizi[4]
    short_name=short_name1[:10]+"-"

    data_tuple = (sector, name, phone, site, address, personels_caount, note, created_date, updated_date, city_id, fount_id, last_status_id, user_id, full_name, reminder, short_name)
    
    cursor.execute("SELECT * FROM crawler_companies WHERE short_name = ? OR phone = ?", (short_name, phone))
    existing_record = cursor.fetchone()
    if existing_record:
##        Eşleşen kayıt varsa, bir sonraki kayda geç
        gecilen=gecilen+1
        ws.append(data_tuple)
        
    else:
        eklenen=eklenen+1
        cursor.execute("INSERT INTO crawler_companies (sector, name, phone, site, address, personels_caount, note, created_date, updated_date, city_id, fount_id, last_status_id, user_id, full_name, reminder, short_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", data_tuple)        
            
    print("Yeni eklenen Kayıt sayısı:",eklenen)
    print("Eklenmeyen kayıt sayısı:",gecilen)
    dizi.clear()
    wb.save("A.xlsx")
    

    
while True:
    firma_cek(firma_no)    
    print(firma_no)

    firma_no=firma_no-1
    connection.commit()
    
     

