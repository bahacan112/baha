from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
##import mysql.connector
from helium import *
import socket
import datetime
import logging
import sys
import sqlite3
# MySQL bağlantısı
#connection = mysql.connector.connect(
#    host='192.168.1.111',
#    port=3306,
#    user='root',
#    password='',
#    database='anonim'
    
#)
connection = sqlite3.connect("db.sqlite3")


firma_no = sys.argv[1]

    

# Genel ilk ayarlar
dizi = []
####wb = load_workbook("1.xlsx")
####ws = wb.active
logging.basicConfig(filename=str(firma_no), level=logging.DEBUG,
                            format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger()
logger.setLevel(logging.INFO)

driver = start_chrome("https://bilgibankasi.ito.org.tr/tr/bilgi-bankasi/firma-bilgileri")  # Chromedriver yolunu güncelleyin

#socket bağlantısı


    



def istemci(c):
    SERVER_IP = '192.168.1.111'
    SERVER_PORT = 12345
    client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    # Sunucuya bağlan
    client_socket.connect((SERVER_IP, SERVER_PORT))
 
   
  
    
    # Veriyi gönder
    data = str(c)+"    "+str(c)
    client_socket.send(data.encode())

    # Sunucudan gelen cevabı al
    response = client_socket.recv(1024)
    print(f"Gelen cevap: {response.decode()}")

# İstemci soketini kapat
    client_socket.close()


# Kod işleme başlangıcı
def firma_cek(company_code):
    a=1
    company_code =int(company_code)
    while True:
          
        try:
            cursor = connection.cursor()
            doubleclick("Ticaret Sicil No")
            press(DELETE)

            write(company_code, into="Ticaret Sicil No")

            click("ara")
            time.sleep(2)
            try:
                secici="/html/body/div[3]/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[2]/td[2]"
                firma=driver.find_element(By.XPATH,secici)
                click(firma)
                time.sleep(2)
            except NoSuchElementException:
       
                 secici="/html/body/div[3]/div/div/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div[2]/div/div[2]/div/div[1]/table/tbody/tr[1]/td[2]"
                 firma=driver.find_element(By.CSS_SELECTOR,secici)
                 click(firma)

            time.sleep(2)
            elements = driver.find_elements(By.CLASS_NAME, "data-cell.col-md-9")

            for element in elements:
                value = element.text
                dizi.append(value[:100])

            ## girilecek veriyi databaseye uygun düzenle
            print(dizi)
            phone = dizi[6]
            element2 = dizi[0]
            print(phone)
#            if len(phone) > 0 and element2 == "":
            if element2 == "":
                phone = phone[:10]  # İlk 10 karakteri al
                # Diğer işlemleri burada gerçekleştirin
                print("Phone:", phone)

                sector = dizi[13]
                name = dizi[4]
                site = dizi[8]
                if site == "":
                    site = " "
                address = dizi[5]
                personels_caount = dizi[1]
                personels_caount = personels_caount.split("-")[0]  # "484725-5" içindeki '-' karakterini ayır ve ilk bölümü al
                personels_caount = ''.join(filter(str.isdigit, personels_caount))  # Yalnızca rakamları tut
                personels_caount = int(personels_caount)
                note = "--kuruluş :" + dizi[9] + "--sermaye :" + dizi[12] + "--açıklama :"
                created_date = "2023-08-20 23:15:54.575094"
                updated_date = "2023-08-20 23:15:54.575094"
                city_id = "61"
                fount_id = int("1")
                last_status_id = int("1")
                user_id = int("1")
                full_name = "DEFAULT"
                reminder = "DEFAULT"
                short_name1 = dizi[4]
                short_name = short_name1[:10] + "-"

                data_tuple = (
                    sector,
                    name,
                    phone,
                    site,
                    address,
                    personels_caount,
                    note,
                    created_date,
                    updated_date,
                    city_id,
                    fount_id,
                    last_status_id,
                    user_id,
                    full_name,
                    None,
                    short_name,
                )

                ## databaseye yazdırma ve testler
####                cursor.execute(
####                    "SELECT * FROM crawler_companies WHERE short_name = %s OR phone = %s",
####                    (short_name, phone),
####                )
####                existing_record = cursor.fetchone()
    
####                if existing_record:
####                    ## Eşleşen kayıt varsa, bir sonraki kayda geç
####                    company_code = company_code + 1
####                    print("yenilenen kayıt")
####                    dizi.clear()
                    
                    
####                else:
                cursor.execute(
                        "INSERT INTO crawler_companies (sector, name, phone, site, address, personels_caount, note, created_date, updated_date, city_id, fount_id, last_status_id, user_id, full_name, reminder, short_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        data_tuple
                )
                    ## databaseye kaydet
                connection.commit()
                logging.info("Veritabanına veri eklendi: %s %s", company_code, "başarılı")

                print("Başarılı")
                dizi.clear()
                print(company_code)
            
                company_code = company_code - 1
####                ws["D1"].value = str(company_code)
####                wb.save("1.xlsx")
                    
                
            else:
                dizi.clear()
                company_code = company_code - 1
                print("Boş Telefon numarası  :", company_code)
                logging.info("Veritabanına veri eklenirken bir hata oluştu:%s %s",company_code,"boş telefon numarası")

        except Exception as err:
            print("Hata:", err)
            logging.info("Veritabanına veri eklenirken bir hata oluştu:%s %s",company_code,"Hata")

            dizi.clear()
            a=a+1
            if a>3:
                    a=1
                    
            continue
            
        finally:
            
            if a==1:
                a=0
                company_code = company_code - 1  
                ##istemci(company_code)
            
            continue

            
        # Bağlantıyı kapat
##        if connection.is_connected():
##            connection.close()
      
    

firma_cek(firma_no)
logging.shutdown()
connection.close()
